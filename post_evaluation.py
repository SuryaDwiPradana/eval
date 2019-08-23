import openpyxl
import json
import os
import shutil

def opensheet(xlsx):
    load = openpyxl.load_workbook(xlsx)
    return load;

def opensheetdata(xlsx):
    load = openpyxl.load_workbook(xlsx, data_only = True)
    return load;

def importJson(txt):
    with open(txt) as f:
        dataJson = json.load(f)
    return dataJson;

def createdir(dirname, path):
    TESTDIR = dirname
    try:
        home = os.path.expanduser(path)

        if not os.path.exists(os.path.join(home, TESTDIR)):  
            os.makedirs(os.path.join(home, TESTDIR))
    except Exception as e:
        print(e)
    return;

def post_main():
	data = importJson('datalist.txt')
	dirname = input("Nama Folder (Post Evaluasi) : ") or "Post Evaluasi"
	path = os.getcwd()

	createdir(dirname,path)
	print("Tolong tunggu sebentar...")
	wb_recap = opensheet('format/Format_2.xlsx')
	counter = 0
	for x in data:
		for y in data[x]:
			for z in y:
				wb_recap.copy_worksheet(wb_recap.active)
				counter +=1
				wb_recap.active = counter
				wb_recap.active.title = z
				wb_recap.active["B1"].value = data[x][0][z][0]['nama']
				wb_recap.active["B2"].value = z
				wb_recap.active["B3"].value = data[x][0][z][0]['jabatan']
				wb_recap.active = 0        

	initTemp = 5
	for x in data:
		for y in data[x]:
			for z in y:
				if os.path.isfile(path+'/Data Evaluasi/'+z+'.xlsx'):
					wb = opensheetdata("Data Evaluasi/"+z+".xlsx")
					for row in range(initTemp,wb.active.max_row,7):
						if wb.active["C"+str(row)].value != z and wb.active["C"+str(row)].value != None:
							wb_recap.active = wb_recap[wb.active["C"+str(row)].value]
							init = 3
							while wb_recap.active["E"+str(init)].value != None:
								init+=1
							wb_recap.active["E"+str(init)].value = data[x][0][z][0]['nama']
							wb_recap.active["F"+str(init)].value = z
							wb_recap.active["G"+str(init)].value = data[x][0][z][0]['jabatan']
							wb_recap.active["H"+str(init)].value = wb.active["F"+str(row+5)].value
							wb_recap.active["I"+str(init)].value = wb.active["G"+str(row+5)].value
							wb_recap.active["J"+str(init)].value = wb.active["H"+str(row+5)].value
							wb_recap.active["K"+str(init)].value = wb.active["K"+str(row+5)].value
							wb_recap.active["L"+str(init)].value = wb.active["N"+str(row+5)].value
							wb_recap.active["M"+str(init)].value = wb.active["S"+str(row+5)].value
							wb_recap.active["N"+str(init)].value = wb.active["W"+str(row+5)].value
							wb_recap.active["O"+str(init)].value = wb.active["AA"+str(row+5)].value
	wb_recap.active = 0
	wb_recap[wb_recap.active.title].sheet_state = "hidden"
	wb_recap.save(path+'/'+dirname+"/Rekap Evaluasi.xlsx")
	return;

#post_main()
