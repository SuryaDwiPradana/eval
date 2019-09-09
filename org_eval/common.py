class Configuration():
    def __init__(self):
        import os
        self.config = {}
        self.folder = {}
        self.report = {}
        self.config['root_path'] = os.getcwd()
        self.config['resources_path'] = os.path.join('org_eval', 'resources')

        self.folder['dirname'] = ['prepare', 'process', 'report']
        self.data = {}
        # self._load_data("datalist.txt")
        self._load_data_new()

    def _load_data_new(self):
        import openpyxl
        file = openpyxl.load_workbook(self.config['resources_path']+"\\"+'data.xlsx')
        data = {}
        file.active = 0
        if file.active.title.lower() == 'data':
            for _ in range(2, file.active.max_row+1):
                name = file.active[f"A{_}"].value
                nim = file.active[f"B{_}"].value
                position = file.active[f"C{_}"].value
                cat = file.active[f"D{_}"].value
                if None in (name, nim, position, cat):
                    continue
                else:
                    name = name.strip()
                    nim = str(nim).strip()
                    position = position.strip()
                    cat = cat.strip()
                if not cat in data:
                    data[cat] = {}
                if not nim in data[cat]:
                    data[cat][nim] = {}
                if not name in data[cat][nim]:
                    data[cat][nim] = {'name': name, 'position': position}
        file.close()
        self.data = data

    def _load_data(self, name):
        try:
            import json
            with open(self.config['resources_path']+"\\"+name) as f:
                self.data = json.load(f)
        except Exception as e:
            print(e)   

    def load_sheet(self, xlsx, data=False):
        import openpyxl
        load = openpyxl.load_workbook(xlsx, data_only=True) if \
            data else openpyxl.load_workbook(xlsx)
        return load

    def load_report(self):
        self.report['period'] = '2018/2019'
        self.report['organization'] = 'BPMU'


def make_dir(dirname, path):
    TESTDIR = dirname
    try:
        import os
        home = os.path.expanduser(path)

        if not os.path.exists(os.path.join(home, TESTDIR)):  
            os.makedirs(os.path.join(home, TESTDIR))
    except Exception as e:
        print(e)

class DataPrepare():
    def __init__(self, excel_name='data'):
        self.dataAll = self.__importExcel(excel_name, 0).dropna()
        self.dataFiltered = self.__filter(self.dataAll)
        self.evaluator = self.__importExcel(excel_name, 1).fillna('')
        self.evalDict = self.__generateDict()

    @staticmethod
    def __importExcel(name, sheet):
        import pandas as pd
        return pd.read_excel('resources/{}.xlsx'.format(name), header = 0, sheet_name = sheet)

    @staticmethod    
    def __filter(data):
        return data.loc[data['Keterangan'] == 'Dievaluasi']

    def __generateDict(self):
        tempDict = {}
        for i in range(len(self.evaluator.Evaluasi)):
            tempDict[self.evaluator.Evaluasi[i]] = [x for x in self.evaluator.iloc[i].values[1:] if x]
        return tempDict

    def gen(self):
        # print(self.dataFiltered)
        # for i in range(len(self.dataFiltered)):
            # print((self.dataFiltered.iloc[i]))
        # x = ((self.dataFiltered.loc[self.dataFiltered['Jabatan'] == 'Anggota Komisi Organisasi']))
        # print(len(x))
        for i in self.evalDict:
            # print(self.evalDict[i])
            # print(i)
            self.genExcel(self.evalDict[i])
            print('------')
            pass
        # pass

    def genExcel(self, jabatan):
        # print(type(jabatan))
        # print(jabatan)
        from openpyxl import load_workbook
        x = load_workbook('resources/format/prepare.xlsx')
        for item in jabatan:
            print(item)
            print('x')
            test = self.dataFiltered.loc[self.dataFiltered['Jabatan'] == item]
            # print(test)
            # test = self.dataFiltered.loc[self.dataFiltered['Jabatan'] in jabatan]
            # print(len(test))
            for i in range(len(test)):
                print(test.Nama.values)
            # for i in range(len(test)):
                x.active['B1'].value = test.Nama.values[i]
                x.active['G1'].value = test.NIM.values[i]
                x.active['J1'].value = test.Jabatan.values[i]
                x.active['N1'].value = test.Nama.values[i]
                x.active['S1'].value = test.Nama.values[i]
            #     if i != len(test):
                x.copy_worksheet(x.active)
            #         x.active = i+1
            #     x.active.title = str(test.NIM.values[i])
            #     # print(test.Nama.values[i])
            #     print(x.active['B1'].value)

        # print(jabatan)
            # pass
        # print(x.active.title)
        # x.active.title = '1'
        # print(x.active.title)
        # print(dir(x.active))
        # x.copy_worksheet(x.active)
        # print(x.sheetnames)
        # print(x.active['A1'].value)
        # print(dir(x))
        x.save("resources/output/prepare/{}.xlsx".format(jabatan))
        x.close()

    def jabatan(self, name):
        print(self.data.loc[(self.data['Keterangan'] == 'Dievaluasi') & (self.data['Jabatan'] == name)])
    
    def report(self):
        pass
x = DataPrepare()
x.gen()
# x.genExcel(["Anggota Komisi Organisasi","Ketua Umum"])
# print(x.data)
# print(x.evaluator)
# print(x.count)
# x.jabatan('Ketua Umum')
# import pprint
# pprint.pprint(x.mapeval)
