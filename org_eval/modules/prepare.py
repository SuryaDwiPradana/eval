class Prepare():
    def __init__(self, excel_name='data'):
        self.__all = self.importExcel(excel_name, 0).dropna()
        self.__filter = self.filter(self.__all)
        self.__evaluator = self.importExcel(excel_name, 1).fillna('')
        self.__evalDict = self.__generateDict()
        self.generate()

    @staticmethod
    def importExcel(filename, sheetname):
        import pandas as pd
        return pd.read_excel('resources/{}.xlsx'.format(filename), header=0, sheet_name=sheetname)

    @staticmethod
    def filter(data):
        return data.loc[data['Keterangan'] == 'Dievaluasi']

    def __generateDict(self):
        tempDict = {}
        for i in range(len(self.__evaluator.Evaluasi)):
            tempDict[self.__evaluator.Evaluasi[i]] = [x for x in self.__evaluator.iloc[i].values[1:] if x]
        return tempDict

    def generate(self):
        for i in self.__evalDict:
            self.__generateExcel(i, self.__evalDict[i])

    def __generateExcel(self, *args):
        from openpyxl import load_workbook
        from openpyxl.styles import Protection
        from openpyxl.workbook.protection import WorkbookProtection
        workbook = load_workbook('resources/format/prepare.xlsx')
        for item in args[1]:
            data = self.__filter.loc[self.__filter['Jabatan'] == item]
            for i in range(len(data)):
                workbook.active.protection.sheet = True
                workbook.active.protection.password = 'sweet'
                for x in range(3, 35):
                    workbook.active['G' + str(x)].protection = Protection(locked=False)
                    workbook.active['L' + str(x)].protection = Protection(locked=False)
                    workbook.active['Q' + str(x)].protection = Protection(locked=False)
                workbook.active['B1'].value = data.Nama.values[i]
                workbook.active['G1'].value = data.NIM.values[i]
                workbook.active['J1'].value = data.Jabatan.values[i]
                workbook.active['N1'].value = data.Nama.values[i]
                workbook.active['S1'].value = data.Nama.values[i]
                workbook.active.title = str(data.NIM.values[i])
                workbook.copy_worksheet(workbook.active)
                workbook.active = len(workbook.sheetnames) - 1
        workbook.remove(workbook.active)
        workbook.security = WorkbookProtection(workbookPassword='sweet', lockWindows=True, lockStructure=True)
        workbook.save("resources/output/prepare/{}.xlsx".format(args[0]))
        workbook.close()
