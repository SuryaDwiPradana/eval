def fill_identity(wb_res, name, nim, org, position, period):
    wb_res.active["B4"] = f": {name}"
    wb_res.active["B5"] = f": {nim}"
    wb_res.active["B6"] = f": {org}"
    wb_res.active["B7"] = f": {position}"
    wb_res.active["B8"] = f": {period}"
    if org.upper().startswith("BPMU") or org.upper().startswith("SMU"):
        wb_res.active["G31"] = f"Ketua Komisi Organisasi {org}"
        wb_res.active["C31"] = f"Ketua Umum {org} UKSW"
    elif org.upper().startswith("BPMF") or org.upper().startswith("SMF"):
        wb_res.active["G31"] = f"Ketua Komisi Organisasi {org}"
        wb_res.active["C31"] = f"Ketua {org} UKSW"


def fill_score(wb_res, wb_ref):
    wb_res.active["F12"] = avgcell(wb_ref, "H")
    wb_res.active["F13"] = avgcell(wb_ref, "I")
    wb_res.active["F14"] = avgcell(wb_ref, "J")
    wb_res.active["F15"] = avgcell(wb_ref, "K")
    wb_res.active["F16"] = avgcell(wb_ref, "L")
    wb_res.active["F17"] = avgcell(wb_ref, "M")
    wb_res.active["F18"] = avgcell(wb_ref, "N")
    wb_res.active["F19"] = avgcell(wb_ref, "O")
    wb_res.active["C22"] = (avgcell(wb_ref, "H") + avgcell(wb_ref, "I") +
                            avgcell(wb_ref, "J") + avgcell(wb_ref, "K") + avgcell(wb_ref, "L") +
                            avgcell(wb_ref, "M") + avgcell(wb_ref, "N") + avgcell(wb_ref, "O")) / 8


def fill_report(wb_res, wb_ref, name, nim, org, position, period):
    fill_identity(wb_res, name, nim, org, position, period)
    fill_score(wb_res, wb_ref)


def avgcell(work, letter):
    count = 0
    xavg = 0
    xsum = 0
    for x in range(3, work.max_row):
        if work[letter + str(x)].value != None:
            count += 1
    if count > 0:
        for x in range(3, count + 3):
            xsum += work[letter + str(x)].value
        xavg = xsum / count
    return xavg


from openpyxl import load_workbook


def importExcel(filename, sheetname, usecols=None, header=0, nrows=None):
    import pandas as pd
    return pd.read_excel('resources/{}.xlsx'.format(filename), header=header, sheet_name=sheetname, usecols=usecols,
                         nrows=nrows)


def main_report():
    try:
        data = importExcel('data', 0).dropna()
        tempList = []
        file_ref = load_workbook('resources/output/process/Summaries.xlsx', data_only=True)
        allRecap = load_workbook('resources/format/report.xlsx')
        singleRecap = load_workbook('resources/format/report.xlsx')

        for l in data.Jabatan.unique():
            tempList.append(l.split(' ', maxsplit=1)[1]) if len(
                l.split(' ', maxsplit=1)[1].split()) > 1 else tempList.append(l)
        listJabatan = list(set(tempList))

        for i in listJabatan:
            print(data.loc[data.Jabatan.str.contains(i)])

        # from tqdm import tqdm
            # report_idx = 0
            # config.load_report()
            # tqdm.write("Loading Files...")
            # file_ref = load_workbook('resources/output/process/Summaries.xlsx')
            # all_recap = load_workbook('resources/format/report.xlsx')
            # solo_recap = load_workbook('resources/format/report.xlsx')
            # tqdm.write("Make Report Files...")
            # for cat in tqdm(listJabatan, ascii=True, desc="Result", position=1):
            # make_dir(cat, 'resources/output/report/')
            #     for nim in tqdm(config.data[cat], ascii=True, desc=cat, position=0):
            #         file_ref.active = file_ref[nim]
            #         all_recap.copy_worksheet(all_recap.active)
            #         report_idx += 1
            #         all_recap.active = report_idx
            #         all_recap.active.title = nim[:9]
            #
            #         fill_report(all_recap, file_ref[nim], config.data[cat][nim]['name'], nim, config.report['organization'],
            #                     config.data[cat][nim]['position'], config.report['period'])
            #         fill_report(solo_recap, file_ref[nim], config.data[cat][nim]['name'], nim,
            #                     config.report['organization'], config.data[cat][nim]['position'], config.report['period'])
            #
            #         solo_recap.save('resources/output/' + cat + '/' + nim + ".xlsx")
            #         all_recap.active = 0
            # all_recap[all_recap.active.title].sheet_state = "hidden"
            # all_recap.save('resources/output/report/Reports.xlsx')
            # tqdm.write("Close All Files...")
    except Exception as e:
        print(e)
    finally:
        input('\nHappy Evaluation\nBest Regard RSS\n')
    return


main_report()
