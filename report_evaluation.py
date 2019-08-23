import openpyxl
import json
import os
import shutil

def opensheet(xlsx):
    load = openpyxl.load_workbook(xlsx)
    return load;

def opensheetdata(xlsx):
    load = openpyxl.load_workbook(xlsx, data_only = True)
    return load;


def importJson(txt):
    with open(txt) as f:
        dataJson = json.load(f)
    return dataJson;

def createdir(dirname, path):
    TESTDIR = dirname
    try:
        home = os.path.expanduser(path)

        if not os.path.exists(os.path.join(home, TESTDIR)):  
            os.makedirs(os.path.join(home, TESTDIR))
    except Exception as e:
        print(e)
    return;

def fillidentity(work,name,nim,org,pos,th):
    work.active["B4"] = ": "+ name
    work.active["B5"] = ": "+ nim
    work.active["B6"] = ": "+ org
    work.active["B7"] = ": "+ pos
    work.active["B8"] = ": "+ th
    return;
def fillscore(workset,workget):
    workset.active["F12"] = avgcell(workget,"H")
    workset.active["F13"] = avgcell(workget,"I")
    workset.active["F14"] = avgcell(workget,"J")
    workset.active["F15"] = avgcell(workget,"K")
    workset.active["F16"] = avgcell(workget,"L")
    workset.active["F17"] = avgcell(workget,"M")
    workset.active["F18"] = avgcell(workget,"N")
    workset.active["F19"] = avgcell(workget,"O")
    workset.active["C22"] = (avgcell(workget,"H")+avgcell(workget,"I")+avgcell(workget,"J")+avgcell(workget,"K")+avgcell(workget,"L")+avgcell(workget,"M")+avgcell(workget,"N")+avgcell(workget,"O"))/8
    return;

def avgcell(work,letter):
    count = 0
    xavg = 0
    xsum = 0
    for x in range(3,work.max_row):
        if work[letter+str(x)].value != None:
            count +=1
    if count > 0:
        for x in range(3,count+3):
            xsum += work[letter+str(x)].value    
        xavg = xsum/count;
    return xavg;

def report_main():
	data = importJson('datalist.txt')
	dirname = input("Nama Folder (Laporan Evaluasi) : ") or "Laporan Evaluasi"
	path = os.getcwd()
	createdir(dirname,path)
	print("Tolong tunggu sebentar...")
	wb_recap = opensheet('format/Format_3.xlsx')
	wb_solo = opensheet('format/Format_3.xlsx')
	rekapfolder = input("Nama Folder Rekap (Post Evaluasi) : ") or "Post Evaluasi"
	rekapfile = input("Nama File Rekap (Rekap Evaluasi) : ") or "Rekap Evaluasi"
	organisasi = input("Nama Organisasi (BPMU) : ") or "BPMU"
	tahun = input("Periode / Bulan (2018/2019) : ") or "2018/2019"
	wb = opensheetdata(rekapfolder+'/'+rekapfile+'.xlsx')
	init = 0
	for x in data:
		createdir(x,path+'/'+dirname)
		for y in data[x]:
			for z in y:
				wb.active = wb[z]
				wb_recap.active = init
				wb_recap.copy_worksheet(wb_recap.active)
				wb_recap.active.title = z[:9]
				fillidentity(wb_recap,data[x][0][z][0]['nama'],z,organisasi,data[x][0][z][0]['jabatan'],tahun)
				fillscore(wb_recap,wb[z])
				
				fillidentity(wb_solo,data[x][0][z][0]['nama'],z,organisasi,data[x][0][z][0]['jabatan'],tahun)
				fillscore(wb_solo,wb[z])
				wb_solo.save(path+'/'+dirname+"/"+x+'/'+z+".xlsx")
				init += 1       
	wb_recap.save(path+'/'+dirname+"/Recap Laporan Evaluasi.xlsx")
	return;

#report_main()
