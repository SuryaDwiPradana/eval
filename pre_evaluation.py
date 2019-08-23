import openpyxl
import json
import os

def opensheet(xlsx):
    load = openpyxl.load_workbook(xlsx)
    return load;

def opensheetdata(xlsx):
    load = openpyxl.load_workbook(xlsx, data_only = True)
    return load;

def importJson(txt):
    with open(txt) as f:
        dataJson = json.load(f)
    return dataJson;

def excelbph(data,path,dirname):
    wb = opensheet('format/Format_1.xlsx')
    for sheet in wb.sheetnames:
        if sheet != "1 BPH" and sheet != "0 Indikator":
            wb[sheet].sheet_state = "hidden"
    wb.active = 0
    while wb.active.title != "1 BPH":
        for x in range(0,len(wb.sheetnames)):
            wb.active = x
            if wb.active.title == "1 BPH":
                break;
    init = 5
    for x in data:
        for y in data[x]:
            for z in y:    
                if wb.active.title == "1 BPH" and x == "BPH":
                    if wb.active["B"+str(init)].value == None:
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
                else:
                    if wb.active["B"+str(init)].value == None and data[x][0][z][0]['jabatan'][:5] == 'Ketua':
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
    wb.save(path+'/'+dirname+"/Evaluasi BPH.xlsx")
    return;

def excelketua(data,path,dirname,komisi):
    wb = opensheet('format/Format_1.xlsx')
    for sheet in wb.sheetnames:
        if sheet != "2 Ketua Komisi" and sheet != "0 Indikator":
            wb[sheet].sheet_state = "hidden"
    wb.active = 0
    while wb.active.title != "2 Ketua Komisi":
        for x in range(0,len(wb.sheetnames)):
            wb.active = x
            if wb.active.title == "2 Ketua Komisi":
                break;
    init = 5
    for x in data:
        for y in data[x]:
            for z in y:    
                if wb.active.title == "2 Ketua Komisi" and x == "BPH":
                    if wb.active["B"+str(init)].value == None:
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
                elif wb.active.title == "2 Ketua Komisi" and x == komisi:
                    if wb.active["B"+str(init)].value == None:
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
                else:
                    if wb.active["B"+str(init)].value == None and data[x][0][z][0]['jabatan'][:5] == 'Ketua':
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
    wb.save(path+'/'+dirname+"/Evaluasi Ketua Komisi "+komisi[-1:]+".xlsx")
    return;

def excelanggota(data,path,dirname,komisi):
    wb = opensheet('format/Format_1.xlsx')
    for sheet in wb.sheetnames:
        if sheet != "3 Anggota Komisi" and sheet != "0 Indikator":
            wb[sheet].sheet_state = "hidden"
    wb.active = 0
    while wb.active.title != "3 Anggota Komisi":
        for x in range(0,len(wb.sheetnames)):
            wb.active = x
            if wb.active.title == "3 Anggota Komisi":
                break;
    init = 5
    for x in data:
        for y in data[x]:
            for z in y:
                if wb.active.title == "3 Anggota Komisi" and x == komisi:
                    if wb.active["B"+str(init)].value == None:
                        wb.active["B"+str(init)].value = data[x][0][z][0]['nama']
                        wb.active["C"+str(init)].value = z
                        wb.active["D"+str(init)].value = data[x][0][z][0]['jabatan']
                        init+=7
    wb.save(path+'/'+dirname+"/Evaluasi Anggota Komisi "+komisi[-1:]+".xlsx")
    return;

def createdir(dirname,path):
    TESTDIR = dirname
    try:
        home = os.path.expanduser(path)

        if not os.path.exists(os.path.join(home, TESTDIR)):  
            os.makedirs(os.path.join(home, TESTDIR))
    except Exception as e:
        print(e)
    return;
	
def pre_main():
	data = importJson('datalist.txt')
	dirname = input("Nama Folder (Pra Evaluasi) : ") or "Pra Evaluasi"
	path = os.getcwd()

	createdir(dirname,path)
	createdir("Data Evaluasi",path)

	print("Pembuatan File Evaluasi(Excel)")
	print("1. Semua")
	print("2. BPH")
	print("3. Komisi A")
	print("4. Komisi B")
	print("5. Komisi C")
	print("6. Komisi D")
	print("7. Ketua Komisi")
	print("8. Anggota Komisi")
	print("9. Exit")
	pilihan = int(input("Silahkan Pilih: ") or "9")
	if pilihan == 1:
		print("Tolong tunggu sebentar...")
		excelbph(data,path,dirname)
		excelketua(data,path,dirname,"KOM A")
		excelanggota(data,path,dirname,"KOM A")
		excelketua(data,path,dirname,"KOM B")
		excelanggota(data,path,dirname,"KOM B")
		excelketua(data,path,dirname,"KOM C")
		excelanggota(data,path,dirname,"KOM C")
		excelketua(data,path,dirname,"KOM D")
		excelanggota(data,path,dirname,"KOM D")
	elif pilihan == 2:
		print("Tolong tunggu sebentar...")
		excelbph(data,path,dirname)
	elif pilihan == 3:
		print("Tolong tunggu sebentar...")
		excelketua(data,path,dirname,"KOM A")
		excelanggota(data,path,dirname,"KOM A")
	elif pilihan == 4:
		print("Tolong tunggu sebentar...")
		excelketua(data,path,dirname,"KOM B")
		excelanggota(data,path,dirname,"KOM B")
	elif pilihan == 5:
		print("Tolong tunggu sebentar...")
		excelketua(data,path,dirname,"KOM C")
		excelanggota(data,path,dirname,"KOM C")
	elif pilihan == 6:
		print("Tolong tunggu sebentar...")
		excelketua(data,path,dirname,"KOM D")
		excelanggota(data,path,dirname,"KOM D")
	elif pilihan == 7:
		print("Tolong tunggu sebentar...")
		excelketua(data,path,dirname,"KOM A")
		excelketua(data,path,dirname,"KOM B")
		excelketua(data,path,dirname,"KOM C")
		excelketua(data,path,dirname,"KOM D")
	elif pilihan == 8:
		print("Tolong tunggu sebentar...")
		excelanggota(data,path,dirname,"KOM A")
		excelanggota(data,path,dirname,"KOM B")
		excelanggota(data,path,dirname,"KOM C")
		excelanggota(data,path,dirname,"KOM D")
	print("Selamat Evaluasi")
	return;

#pre_main()
